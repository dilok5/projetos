from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook  # type: ignore

# Inicia o navegador
navegador = webdriver.Chrome()

try:
    # Acessa o Google
    navegador.get('https://www.google.com')

    # Localiza a barra de pesquisa e busca por "cotação dolar"
    WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="APjFqb"]')))
    navegador.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys('cotação dólar', Keys.ENTER)

    # Aguarda o elemento do valor do dólar ser carregado
    cotacao_xpath = '//*[@id="knowledge-currency__updatable-data-column"]/div[1]/div[2]/span[1]'
    valor_dolar_elemento = WebDriverWait(navegador, 10).until(
        EC.presence_of_element_located((By.XPATH, cotacao_xpath))
    )

    # Captura o valor do dólar
    valor_dolar = valor_dolar_elemento.text  # Use `.text` para obter o valor exibido
    print(f'Valor atual do dólar: R$ {valor_dolar}')

    # Abre o arquivo Excel
    arquivo = load_workbook('produtos.xlsx')
    plan = arquivo.active

    # Atualiza o valor do dólar na célula correspondente
    plan['B1'] = float(valor_dolar.replace(',', '.'))  # Substitui a vírgula por ponto para converter em float

    # Recalcula a coluna "Total US$"
    for linha in range(5, plan.max_row + 1):  # Começa na linha onde os dados reais iniciam
        try:
            quantidade = plan.cell(row=linha, column=2).value
            preco_unitario = plan.cell(row=linha, column=3).value

            # Ignora linhas inválidas ou cabeçalhos
            if isinstance(quantidade, (int, float)) and isinstance(preco_unitario, str) and "R$" in preco_unitario:
                preco_unitario = float(preco_unitario.replace('R$ ', '').replace(',', '.'))
                total_reais = quantidade * preco_unitario
                total_dolares = total_reais / plan['B1'].value
                plan.cell(row=linha, column=5).value = f'R$ {total_dolares:.2f}'
        except Exception as linha_erro:
            print(f"Erro ao processar a linha {linha}: {linha_erro}")

    # Salva o arquivo Excel
    arquivo.save('produtos_atualizados.xlsx')
    print("Planilha atualizada com sucesso!")

except Exception as e:
    print("Ocorreu um erro:", e)

finally:
    navegador.quit()  # Garante que o navegador será fechado
